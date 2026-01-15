from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List, Any, Optional

from src.core.exceptions import ProcessingError
from src.core.models.office import OfficeTextContent
from ._common import sanitize_filename, ensure_dir


@dataclass(slots=True)
class XlsxExtraction:
    markdown: str
    image_paths: List[str]
    text_contents: List[OfficeTextContent]


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(v)
    return str(v).strip()


def _to_md_table(rows: List[List[str]]) -> str:
    if not rows:
        return ""
    col_count = max(len(r) for r in rows)
    norm = [r + [""] * (col_count - len(r)) for r in rows]
    header = norm[0]
    out: List[str] = []
    out.append("| " + " | ".join(header) + " |")
    out.append("| " + " | ".join(["---"] * col_count) + " |")
    for r in norm[1:]:
        out.append("| " + " | ".join(r) + " |")
    return "\n".join(out)


def extract_xlsx_to_markdown(
    input_path: Path,
    images_dir: Path,
    max_rows: int = 200,
    max_cols: int = 30,
) -> XlsxExtraction:
    """Extrahiert XLSX nach Markdown.

    Fokus: robuste Tabellenwerte. Layout-/Chart-Rendering ist in Pipeline B besser aufgehoben.
    """
    try:
        import openpyxl  # type: ignore
    except Exception as e:  # pragma: no cover
        raise ProcessingError(f"openpyxl ist nicht installiert: {e}")

    ensure_dir(images_dir)
    wb = openpyxl.load_workbook(str(input_path), data_only=True)

    lines: List[str] = []
    image_paths: List[str] = []
    text_contents: List[OfficeTextContent] = []

    for sheet_index, ws in enumerate(wb.worksheets):
        title = ws.title or f"Sheet {sheet_index + 1}"
        lines.append(f"## Sheet: {title}")

        # Tabellenbereich – bewusst begrenzt, damit Output kontrollierbar bleibt.
        rows: List[List[str]] = []
        # iter_rows ist 1-basiert
        max_r = min(ws.max_row or 0, max_rows)
        max_c = min(ws.max_column or 0, max_cols)
        for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c, values_only=True):
            rows.append([_cell_str(v) for v in row])

        # Truncation Hinweis (für Vergleichbarkeit/Debugging)
        if (ws.max_row or 0) > max_rows or (ws.max_column or 0) > max_cols:
            lines.append(
                f"_Hinweis: Tabelle gekürzt auf {max_rows} Zeilen / {max_cols} Spalten (original: {ws.max_row}x{ws.max_column})._"
            )

        if rows:
            lines.append(_to_md_table(rows))
            flat_text = "\n".join(["\t".join(r) for r in rows])
            text_contents.append(OfficeTextContent(index=sheet_index, text=flat_text, kind="xlsx_sheet"))
        else:
            text_contents.append(OfficeTextContent(index=sheet_index, text="", kind="xlsx_sheet"))

        # Embedded Images (best-effort). openpyxl API ist nicht in allen Fällen stabil.
        imgs: List[Any] = list(getattr(ws, "_images", []) or [])
        for img_index, img in enumerate(imgs):
            try:
                blob: Optional[bytes] = None
                if hasattr(img, "_data") and callable(getattr(img, "_data")):
                    blob = img._data()  # type: ignore[attr-defined]
                if not blob:
                    continue
                ext = ".png"
                filename = sanitize_filename(f"xlsx-img-{sheet_index + 1}-{img_index + 1}") + ext
                out_path = images_dir / filename
                with open(out_path, "wb") as f:
                    f.write(blob)
                rel = f"images/{filename}"
                image_paths.append(rel)
                lines.append(f"![]({rel})")
            except Exception:
                continue

        lines.append("")

    markdown = "\n".join(lines).strip() + "\n"
    return XlsxExtraction(markdown=markdown, image_paths=image_paths, text_contents=text_contents)






