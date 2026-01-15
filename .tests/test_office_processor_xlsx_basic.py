from __future__ import annotations

import asyncio
import tempfile
from pathlib import Path

import pytest

from src.processors.office_processor import OfficeProcessor


def _make_xlsx(tmp: Path) -> Path:
    try:
        import openpyxl  # type: ignore
    except Exception as e:  # pragma: no cover
        pytest.skip(f"openpyxl nicht verfügbar: {e}")

    p = tmp / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Wert"
    ws["A2"] = "X"
    ws["B2"] = 42
    wb.save(str(p))
    return p


def test_office_processor_xlsx_creates_markdown() -> None:
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        xlsx_path = _make_xlsx(tmp)

        proc = OfficeProcessor(process_id="test-xlsx")
        res = asyncio.run(
            proc.process(
                xlsx_path,
                include_images=True,
                include_previews=True,
                use_cache=False,
                base_cache_dir=tmp / "cache",
            )
        )

        assert res.data.metadata.format == "xlsx"
        assert "## Sheet:" in res.data.extracted_text
        assert "Name" in res.data.extracted_text
        assert "42" in res.data.extracted_text






